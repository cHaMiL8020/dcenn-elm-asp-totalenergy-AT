import argparse
import itertools
import json
import os
import time

import clingo
import numpy as np
import pandas as pd
import torch
import torch.nn as nn
import torch.optim as optim
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sklearn.metrics import mean_absolute_error, mean_squared_error
from sklearn.preprocessing import StandardScaler


class CentroidEncoder(nn.Module):
    def __init__(self, input_dim, latent_dim):
        super(CentroidEncoder, self).__init__()
        self.encoder = nn.Sequential(
            nn.Linear(input_dim, 32),
            nn.ReLU(),
            nn.Linear(32, 16),
            nn.ReLU(),
            nn.Linear(16, latent_dim),
        )
        self.decoder = nn.Sequential(
            nn.Linear(latent_dim, 16),
            nn.ReLU(),
            nn.Linear(16, 32),
            nn.ReLU(),
            nn.Linear(32, input_dim),
        )

    def forward(self, x):
        latent = self.encoder(x)
        reconstructed = self.decoder(latent)
        return latent, reconstructed


class ELMRegressor:
    def __init__(self, input_dim, hidden_dim):
        self.input_dim = input_dim
        self.hidden_dim = hidden_dim
        self.W = np.random.randn(input_dim, hidden_dim)
        self.b = np.random.randn(hidden_dim)
        self.beta = None

    def _relu(self, x):
        return np.maximum(0, x)

    def fit(self, x_train, y_train):
        h = self._relu(np.dot(x_train, self.W) + self.b)
        h_pinv = np.linalg.pinv(h)
        self.beta = np.dot(h_pinv, y_train)

    def predict(self, x_test):
        h = self._relu(np.dot(x_test, self.W) + self.b)
        return np.dot(h, self.beta)


def run_asp_anomaly_counts(df_subset, rules_path):
    ctl = clingo.Control(["0", "--warn=no-atom-undefined"])
    ctl.load(rules_path)

    facts = []
    for idx, row in enumerate(df_subset.itertuples()):
        pred_val = int(row.predicted_generation)
        actual_val = int(row.power_generation)
        cglo_val = int(row.cglo)
        facts.append(f"pred({idx}, {pred_val}).")
        facts.append(f"actual({idx}, {actual_val}).")
        facts.append(f"cglo({idx}, {cglo_val}).")

    ctl.add("base", [], "\n".join(facts))
    ctl.ground([("base", [])])

    anomalies = []

    def on_model(model):
        for symbol in model.symbols(atoms=True):
            if symbol.name == "anomaly":
                anomalies.append(
                    {
                        "time_idx": symbol.arguments[0].number,
                        "reason": symbol.arguments[1].string,
                    }
                )

    ctl.solve(on_model=on_model)

    if not anomalies:
        return {
            "asp_anomaly_count": 0,
            "asp_critical_deviation_count": 0,
            "asp_ramp_spike_count": 0,
            "asp_ramp_drop_count": 0,
            "asp_negative_gen_count": 0,
            "asp_below_baseline_count": 0,
        }

    anomaly_df = pd.DataFrame(anomalies)
    reasons = anomaly_df["reason"].value_counts()

    return {
        "asp_anomaly_count": int(len(anomaly_df)),
        "asp_critical_deviation_count": int(reasons.get("Critical AI Deviation", 0)),
        "asp_ramp_spike_count": int(reasons.get("Ramp Rate Violation (Spike)", 0)),
        "asp_ramp_drop_count": int(reasons.get("Ramp Rate Violation (Drop)", 0)),
        "asp_negative_gen_count": int(reasons.get("Negative Generation Predicted", 0)),
        "asp_below_baseline_count": int(reasons.get("Predicted Below Historical Grid Baseline", 0)),
    }


def build_feature_sets(available_columns):
    feature_sets = {
        "baseline": [
            "cglo",
            "ffam",
            "rr",
            "tl",
            "time_sin",
            "time_cos",
            "month_sin",
            "month_cos",
            "power_lag_24h",
        ],
        "baseline_plus_calendar": [
            "cglo",
            "ffam",
            "rr",
            "tl",
            "time_sin",
            "time_cos",
            "month_sin",
            "month_cos",
            "power_lag_24h",
            "hour",
            "day_of_week",
        ],
        "baseline_plus_minute": [
            "cglo",
            "ffam",
            "rr",
            "tl",
            "time_sin",
            "time_cos",
            "month_sin",
            "month_cos",
            "power_lag_24h",
            "hour",
            "minute",
            "day_of_week",
        ],
        "weather_plus_time": [
            "cglo",
            "ffam",
            "rr",
            "tl",
            "hour",
            "minute",
            "month",
            "day_of_week",
            "time_sin",
            "time_cos",
            "month_sin",
            "month_cos",
        ],
    }

    filtered_sets = {}
    for name, cols in feature_sets.items():
        present_cols = [col for col in cols if col in available_columns]
        if present_cols:
            filtered_sets[name] = present_cols

    return filtered_sets


def train_and_evaluate(df, config, rules_path, asp_window_size):
    seed = config["seed"]
    np.random.seed(seed)
    torch.manual_seed(seed)

    train_df = df[df["timestamp"] < "2024-01-01"]
    test_df = df[df["timestamp"] >= "2024-01-01"]

    x_train_raw = train_df[config["features"]].values
    y_train = train_df["power_generation"].values
    x_test_raw = test_df[config["features"]].values
    y_test = test_df["power_generation"].values

    scaler = StandardScaler()
    x_train_scaled = scaler.fit_transform(x_train_raw)
    x_test_scaled = scaler.transform(x_test_raw)

    x_train_tensor = torch.FloatTensor(x_train_scaled)
    x_test_tensor = torch.FloatTensor(x_test_scaled)

    autoencoder = CentroidEncoder(x_train_scaled.shape[1], config["latent_dim"])
    criterion = nn.MSELoss()
    optimizer = optim.Adam(autoencoder.parameters(), lr=config["lr"])

    for _ in range(config["epochs"]):
        optimizer.zero_grad()
        _, reconstructed = autoencoder(x_train_tensor)
        loss = criterion(reconstructed, x_train_tensor)
        loss.backward()
        optimizer.step()

    with torch.no_grad():
        z_train, _ = autoencoder(x_train_tensor)
        z_test, _ = autoencoder(x_test_tensor)

    elm = ELMRegressor(input_dim=config["latent_dim"], hidden_dim=config["elm_hidden_neurons"])
    elm.fit(z_train.numpy(), y_train)
    predictions = elm.predict(z_test.numpy())

    rmse = float(np.sqrt(mean_squared_error(y_test, predictions)))
    mae = float(mean_absolute_error(y_test, predictions))

    abs_err = np.abs(y_test - predictions)
    evening_mask = test_df["timestamp"].dt.hour.between(18, 23)
    evening_mae = float(abs_err[evening_mask.values].mean())

    mape = float(np.nanmean(np.abs((y_test - predictions) / np.where(y_test == 0, np.nan, y_test))) * 100)
    mean_actual = float(np.mean(y_test))
    std_actual = float(np.std(y_test))

    result = {
        "feature_set": config["feature_set_name"],
        "feature_count": len(config["features"]),
        "features": "|".join(config["features"]),
        "latent_dim": config["latent_dim"],
        "elm_hidden_neurons": config["elm_hidden_neurons"],
        "epochs": config["epochs"],
        "learning_rate": config["lr"],
        "seed": seed,
        "rmse": rmse,
        "mae": mae,
        "mape_percent": mape,
        "nrmse_mean": float(rmse / mean_actual) if mean_actual else np.nan,
        "nrmse_std": float(rmse / std_actual) if std_actual else np.nan,
        "evening_mae_18_23": evening_mae,
    }

    pred_df = test_df[["timestamp", "power_generation", "cglo"]].copy()
    pred_df["predicted_generation"] = predictions
    asp_df = pred_df.head(asp_window_size).copy()
    asp_metrics = run_asp_anomaly_counts(asp_df, rules_path)
    result.update(asp_metrics)

    return result


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

_HDR_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
_ALT_FILL   = PatternFill("solid", fgColor="EBF3FF")   # pale blue row stripe
_GOLD_FILL  = PatternFill("solid", fgColor="FFD700")   # gold for winner row
_HDR_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_BODY_FONT  = Font(name="Calibri", size=10)
_BOLD_FONT  = Font(name="Calibri", bold=True, size=10)
_THIN       = Side(style="thin", color="B0B0B0")
_BORDER     = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=False)
_WRAP       = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _apply_header_row(ws, row_idx: int, titles: list):
    """Write a navy-coloured header row."""
    for col_idx, title in enumerate(titles, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=title)
        cell.fill   = _HDR_FILL
        cell.font   = _HDR_FONT
        cell.border = _BORDER
        cell.alignment = _CENTER


def _autofit_columns(ws, min_width=10, max_width=40):
    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(max_len + 2, min_width), max_width)


def _write_summary_sheet(wb, summary_df):
    ws = wb.create_sheet("Summary (Aggregated)", 0)

    display_cols = [
        "feature_set", "latent_dim", "elm_hidden_neurons", "epochs", "learning_rate",
        "rmse_mean", "rmse_std", "mae_mean", "mae_std",
        "mape_percent_mean", "evening_mae_18_23_mean",
        "asp_anomaly_count_mean", "nrmse_mean_mean",
    ]
    present = [c for c in display_cols if c in summary_df.columns]

    pretty = {
        "feature_set": "Feature Set",
        "latent_dim": "Latent Dim",
        "elm_hidden_neurons": "ELM Neurons",
        "epochs": "Epochs",
        "learning_rate": "Learning Rate",
        "rmse_mean": "RMSE (mean)",
        "rmse_std": "RMSE (std)",
        "mae_mean": "MAE (mean)",
        "mae_std": "MAE (std)",
        "mape_percent_mean": "MAPE % (mean)",
        "evening_mae_18_23_mean": "Evening MAE 18-23h",
        "asp_anomaly_count_mean": "ASP Anomalies",
        "nrmse_mean_mean": "NRMSE",
    }

    # Title banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(present))
    title_cell = ws.cell(row=1, column=1, value="dCeNN-ELM Benchmark — Aggregated Results (sorted by RMSE)")
    title_cell.fill = PatternFill("solid", fgColor="0D1B2A")
    title_cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    title_cell.alignment = _CENTER
    ws.row_dimensions[1].height = 24

    _apply_header_row(ws, 2, [pretty.get(c, c) for c in present])
    ws.row_dimensions[2].height = 18

    sorted_df = summary_df.sort_values("rmse_mean").reset_index(drop=True)
    fmt_map = {
        "rmse_mean": "0.0000", "rmse_std": "0.0000",
        "mae_mean": "0.0000",  "mae_std": "0.0000",
        "mape_percent_mean": "0.00",
        "evening_mae_18_23_mean": "0.0000",
        "nrmse_mean_mean": "0.0000",
        "learning_rate": "0.0000",
    }

    for r_idx, (_, row) in enumerate(sorted_df.iterrows(), start=3):
        is_best = r_idx == 3
        fill = _GOLD_FILL if is_best else (_ALT_FILL if r_idx % 2 == 0 else PatternFill())
        for c_idx, col in enumerate(present, start=1):
            val = row[col]
            if isinstance(val, float) and pd.isna(val):
                val = ""
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill   = fill
            cell.font   = _BOLD_FONT if is_best else _BODY_FONT
            cell.border = _BORDER
            cell.alignment = _CENTER
            if col in fmt_map and isinstance(val, (int, float)):
                cell.number_format = fmt_map[col]

    # Conditional colour scale on RMSE column
    rmse_col_idx = present.index("rmse_mean") + 1 if "rmse_mean" in present else None
    if rmse_col_idx:
        rmse_col_letter = get_column_letter(rmse_col_idx)
        last_row = 2 + len(sorted_df)
        rule = ColorScaleRule(
            start_type="min", start_color="63BE7B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="F8696B",
        )
        ws.conditional_formatting.add(f"{rmse_col_letter}3:{rmse_col_letter}{last_row}", rule)

    ws.freeze_panes = "A3"
    _autofit_columns(ws)
    ws.sheet_view.showGridLines = False


def _write_runs_sheet(wb, runs_df):
    ws = wb.create_sheet("All Runs (Raw)")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(runs_df.columns))
    title_cell = ws.cell(row=1, column=1, value="All Individual Benchmark Runs — Raw Data")
    title_cell.fill = PatternFill("solid", fgColor="1A3A5C")
    title_cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    title_cell.alignment = _CENTER
    ws.row_dimensions[1].height = 22

    _apply_header_row(ws, 2, list(runs_df.columns))
    ws.row_dimensions[2].height = 18

    numeric_fmt = {
        "rmse": "0.0000", "mae": "0.0000", "mape_percent": "0.00",
        "nrmse_mean": "0.0000", "nrmse_std": "0.0000",
        "evening_mae_18_23": "0.0000", "learning_rate": "0.0000",
        "runtime_seconds": "0.00",
    }

    for r_idx, (_, row) in enumerate(runs_df.iterrows(), start=3):
        fill = _ALT_FILL if r_idx % 2 == 0 else PatternFill()
        for c_idx, col in enumerate(runs_df.columns, start=1):
            val = row[col]
            if isinstance(val, float) and pd.isna(val):
                val = ""
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill      = fill
            cell.font      = _BODY_FONT
            cell.border    = _BORDER
            cell.alignment = _CENTER
            if col in numeric_fmt and isinstance(val, (int, float)):
                cell.number_format = numeric_fmt[col]

    ws.freeze_panes = "A3"
    _autofit_columns(ws)
    ws.sheet_view.showGridLines = False


def _write_best_config_sheet(wb, best_row, total_runtime):
    ws = wb.create_sheet("Best Config")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 35

    ws.merge_cells("A1:B1")
    title_cell = ws.cell(row=1, column=1, value="Best Configuration — Lowest Mean RMSE")
    title_cell.fill = PatternFill("solid", fgColor="0D3B66")
    title_cell.font = Font(name="Calibri", bold=True, color="FFD700", size=13)
    title_cell.alignment = _CENTER
    ws.row_dimensions[1].height = 28

    entries = [
        ("Feature Set",             best_row["feature_set"]),
        ("Feature Count",           int(best_row["feature_count"])),
        ("Latent Dimension",        int(best_row["latent_dim"])),
        ("ELM Hidden Neurons",      int(best_row["elm_hidden_neurons"])),
        ("Epochs",                  int(best_row["epochs"])),
        ("Learning Rate",           float(best_row["learning_rate"])),
        (None, None),
        ("RMSE (mean ± std)",       f"{best_row['rmse_mean']:.4f} ± {best_row.get('rmse_std', 0):.4f}"),
        ("MAE  (mean ± std)",       f"{best_row['mae_mean']:.4f} ± {best_row.get('mae_std', 0):.4f}"),
        ("MAPE %",                  f"{best_row['mape_percent_mean']:.2f}%"),
        ("Evening MAE 18–23h",      f"{best_row['evening_mae_18_23_mean']:.4f}"),
        ("NRMSE",                   f"{best_row['nrmse_mean_mean']:.4f}" if 'nrmse_mean_mean' in best_row else "N/A"),
        ("ASP Anomalies (mean)",    int(best_row["asp_anomaly_count_mean"])),
        (None, None),
        ("Benchmark Total Runtime", f"{total_runtime:.1f} s"),
    ]

    section_header_rows = {2: "Hyperparameters", 8: "Evaluation Metrics", 14: "Runtime"}

    for r_offset, (label, value) in enumerate(entries, start=2):
        if r_offset in section_header_rows:
            ws.merge_cells(start_row=r_offset, start_column=1, end_row=r_offset, end_column=2)
            sec_cell = ws.cell(row=r_offset, column=1, value=section_header_rows[r_offset])
            sec_cell.fill = PatternFill("solid", fgColor="2E75B6")
            sec_cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
            sec_cell.alignment = _CENTER
            continue
        if label is None:
            continue
        lc = ws.cell(row=r_offset, column=1, value=label)
        lc.font = _BOLD_FONT
        lc.border = _BORDER
        lc.fill = PatternFill("solid", fgColor="DEEAF1")
        lc.alignment = Alignment(horizontal="left", vertical="center")

        vc = ws.cell(row=r_offset, column=2, value=value)
        vc.font = _BODY_FONT
        vc.border = _BORDER
        vc.alignment = _CENTER

    ws.sheet_view.showGridLines = False


def write_excel_report(runs_df, summary_df, best_row, total_runtime, out_xlsx_path):
    """Write a multi-sheet, formatted Excel workbook with all benchmark results."""
    from openpyxl import Workbook
    wb = Workbook()
    # Remove default blank sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _write_summary_sheet(wb, summary_df)
    _write_runs_sheet(wb, runs_df)
    _write_best_config_sheet(wb, best_row, total_runtime)

    wb.save(out_xlsx_path)
    print(f"Saved formatted Excel workbook to: {out_xlsx_path}")


# ---------------------------------------------------------------------------
# Legacy plain-text report (kept for git-diffable reference)
# ---------------------------------------------------------------------------

def write_report(summary_df, best_row, out_md_path):
    lines = []
    lines.append("# dCeNN-ELM Benchmark Report")
    lines.append("")
    lines.append("## Best Configuration")
    lines.append("")
    lines.append(f"- Feature set: {best_row['feature_set']}")
    lines.append(f"- Feature count: {int(best_row['feature_count'])}")
    lines.append(f"- Latent dim: {int(best_row['latent_dim'])}")
    lines.append(f"- ELM hidden neurons: {int(best_row['elm_hidden_neurons'])}")
    lines.append(f"- Epochs: {int(best_row['epochs'])}")
    lines.append(f"- Learning rate: {best_row['learning_rate']}")
    lines.append(f"- RMSE: {best_row['rmse_mean']:.4f}")
    lines.append(f"- MAE: {best_row['mae_mean']:.4f}")
    lines.append(f"- MAPE: {best_row['mape_percent_mean']:.2f}%")
    lines.append(f"- Evening MAE (18-23): {best_row['evening_mae_18_23_mean']:.4f}")
    lines.append(f"- ASP anomalies in test window: {int(best_row['asp_anomaly_count_mean'])}")
    lines.append("")
    lines.append("## Top 10 Configurations by RMSE")
    lines.append("")

    top10 = summary_df.sort_values("rmse_mean").head(10)
    lines.append("| feature_set | latent_dim | elm_hidden_neurons | epochs | lr | rmse | mae | mape% | asp_anomalies |")
    lines.append("|---|---:|---:|---:|---:|---:|---:|---:|---:|")
    for _, row in top10.iterrows():
        lines.append(
            "| "
            + f"{row['feature_set']} | {int(row['latent_dim'])} | {int(row['elm_hidden_neurons'])} | "
            + f"{int(row['epochs'])} | {row['learning_rate']} | {row['rmse_mean']:.4f} | "
            + f"{row['mae_mean']:.4f} | {row['mape_percent_mean']:.2f} | {int(row['asp_anomaly_count_mean'])} |"
        )

    with open(out_md_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")


def main():
    parser = argparse.ArgumentParser(description="Benchmark dCeNN-ELM configurations.")
    parser.add_argument("--processed-path", default="data/processed_15min.csv")
    parser.add_argument("--rules-path", default="rules/grid_rules.lp")
    parser.add_argument("--output-dir", default="data/benchmarks")
    parser.add_argument("--asp-window-size", type=int, default=2976)
    parser.add_argument("--quick", action="store_true", help="Run a smaller benchmark sweep.")
    args = parser.parse_args()

    if not os.path.exists(args.processed_path):
        raise FileNotFoundError(f"Missing processed dataset: {args.processed_path}")
    if not os.path.exists(args.rules_path):
        raise FileNotFoundError(f"Missing ASP rules file: {args.rules_path}")

    os.makedirs(args.output_dir, exist_ok=True)

    print("Loading processed dataset...")
    df = pd.read_csv(args.processed_path)
    df["timestamp"] = pd.to_datetime(df["timestamp"])

    feature_sets = build_feature_sets(df.columns)
    if not feature_sets:
        raise ValueError("No valid feature sets found in processed dataset columns.")

    if args.quick:
        grid = {
            "latent_dim": [8, 10],
            "elm_hidden_neurons": [128, 256],
            "epochs": [50],
            "lr": [0.01],
            "seed": [42],
        }
    else:
        grid = {
            "latent_dim": [8, 10],
            "elm_hidden_neurons": [100, 256],
            "epochs": [50, 120],
            "lr": [0.01, 0.005],
            "seed": [42, 1337],
        }

    keys = list(grid.keys())
    run_configs = []
    for feature_set_name, feature_list in feature_sets.items():
        for values in itertools.product(*[grid[k] for k in keys]):
            config = dict(zip(keys, values))
            config["feature_set_name"] = feature_set_name
            config["features"] = feature_list
            run_configs.append(config)

    print(f"Running benchmark for {len(run_configs)} configurations...")

    rows = []
    start = time.time()

    for idx, config in enumerate(run_configs, start=1):
        run_start = time.time()
        result = train_and_evaluate(df, config, args.rules_path, args.asp_window_size)
        result["runtime_seconds"] = round(time.time() - run_start, 3)
        rows.append(result)
        print(
            f"[{idx}/{len(run_configs)}] {result['feature_set']} ld={result['latent_dim']} "
            + f"h={result['elm_hidden_neurons']} ep={result['epochs']} lr={result['learning_rate']} "
            + f"seed={result['seed']} -> RMSE={result['rmse']:.4f}, MAE={result['mae']:.4f}, "
            + f"ASP={result['asp_anomaly_count']}"
        )

    total_runtime = round(time.time() - start, 3)

    runs_df = pd.DataFrame(rows)
    runs_csv = os.path.join(args.output_dir, "dcenn_elm_benchmark_runs.csv")
    runs_df.to_csv(runs_csv, index=False)

    group_cols = [
        "feature_set",
        "feature_count",
        "features",
        "latent_dim",
        "elm_hidden_neurons",
        "epochs",
        "learning_rate",
    ]
    metrics = [
        "rmse",
        "mae",
        "mape_percent",
        "nrmse_mean",
        "nrmse_std",
        "evening_mae_18_23",
        "asp_anomaly_count",
        "asp_critical_deviation_count",
        "asp_ramp_spike_count",
        "asp_ramp_drop_count",
        "asp_negative_gen_count",
        "asp_below_baseline_count",
        "runtime_seconds",
    ]

    agg_map = {m: ["mean", "std"] for m in metrics}
    summary_df = runs_df.groupby(group_cols, dropna=False).agg(agg_map).reset_index()
    flattened_cols = []
    for col in summary_df.columns.to_flat_index():
        if isinstance(col, str):
            flattened_cols.append(col)
        elif col[1] == "":
            flattened_cols.append(col[0])
        else:
            flattened_cols.append(f"{col[0]}_{col[1]}")
    summary_df.columns = flattened_cols

    summary_csv = os.path.join(args.output_dir, "dcenn_elm_benchmark_summary.csv")
    summary_df.sort_values("rmse_mean").to_csv(summary_csv, index=False)

    best_row = summary_df.sort_values("rmse_mean").iloc[0]
    best_config = {
        "feature_set": best_row["feature_set"],
        "feature_count": int(best_row["feature_count"]),
        "features": best_row["features"].split("|"),
        "latent_dim": int(best_row["latent_dim"]),
        "elm_hidden_neurons": int(best_row["elm_hidden_neurons"]),
        "epochs": int(best_row["epochs"]),
        "learning_rate": float(best_row["learning_rate"]),
        "rmse_mean": float(best_row["rmse_mean"]),
        "mae_mean": float(best_row["mae_mean"]),
        "mape_percent_mean": float(best_row["mape_percent_mean"]),
        "asp_anomaly_count_mean": float(best_row["asp_anomaly_count_mean"]),
        "benchmark_total_runtime_seconds": total_runtime,
    }

    best_json = os.path.join(args.output_dir, "dcenn_elm_best_config.json")
    with open(best_json, "w", encoding="utf-8") as f:
        json.dump(best_config, f, indent=2)

    report_md = os.path.join(args.output_dir, "dcenn_elm_benchmark_report.md")
    write_report(summary_df, best_row, report_md)

    report_xlsx = os.path.join(args.output_dir, "dcenn_elm_benchmark_report.xlsx")
    write_excel_report(runs_df, summary_df, best_row, total_runtime, report_xlsx)

    print("Benchmark complete.")
    print(f"Saved detailed runs to: {runs_csv}")
    print(f"Saved summarized benchmarks to: {summary_csv}")
    print(f"Saved best configuration to: {best_json}")
    print(f"Saved researcher report (markdown) to: {report_md}")
    print(f"Saved formatted Excel workbook to: {report_xlsx}")
    print(f"Total benchmark runtime (s): {total_runtime}")


if __name__ == "__main__":
    main()
